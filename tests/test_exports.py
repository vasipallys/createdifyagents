"""Final-result workbook and download endpoint coverage."""

import io

from fastapi.testclient import TestClient
from openpyxl import load_workbook

import story_pointer.api as api_module
from story_pointer.exports import ExportResultItem, build_results_workbook
from story_pointer.schema import StoryPointResult


def _result(title: str = "Payment change") -> StoryPointResult:
    return StoryPointResult(
        ok=True,
        title=title,
        points=5,
        plain_language_why="Cross-layer work with bounded integration.",
        tldr="Moderate, well-bounded delivery.",
        factors=[{"id": "integration_surface", "level": "Medium", "evidence": "One API"}],
        risks=[{"description": "Vendor delay", "severity": "Medium", "mitigation": "Mock first"}],
        hidden_work=["Regression tests"],
        assumptions=["Existing API remains compatible"],
        recommended_split=[{"title": "API", "points": 3, "why": "Separate backend work"}],
        person_days={"min": 4, "max": 6},
        provider="openai",
        model="gpt-4o-mini",
    )


def test_workbook_contains_summary_and_normalized_detail_sheets():
    content = build_results_workbook([
        ExportResultItem(index=0, title="Payment change", status="done", result=_result()),
        ExportResultItem(index=1, title="Failed story", status="failed", error="Provider unavailable"),
    ])

    workbook = load_workbook(io.BytesIO(content), data_only=False)
    assert workbook.sheetnames == ["Summary", "Factors", "Risks", "Supporting detail"]
    summary = workbook["Summary"]
    assert summary.freeze_panes == "A2"
    assert summary["B2"].value == "Payment change"
    assert summary["C2"].value == "Certified"
    assert summary["D2"].value == 5
    assert summary["C3"].value == "Failed / redacted"
    assert summary["M3"].value == "Provider unavailable"
    assert workbook["Factors"].max_row == 2
    assert workbook["Risks"].max_row == 2
    assert workbook["Supporting detail"].max_row >= 5


def test_workbook_neutralizes_formula_like_text():
    content = build_results_workbook([
        ExportResultItem(index=0, title="=HYPERLINK(\"bad\")", status="done", result=_result()),
    ])
    workbook = load_workbook(io.BytesIO(content), data_only=False)

    assert workbook["Summary"]["B2"].value.startswith("'=")
    assert workbook["Summary"]["B2"].data_type != "f"


def test_excel_export_endpoint_returns_real_xlsx():
    payload = {
        "items": [{
            "index": 0,
            "title": "Payment change",
            "status": "done",
            "result": _result().model_dump(),
            "error": "",
            "trace_id": "trace-123",
        }]
    }
    with TestClient(api_module.app) as client:
        response = client.post("/export/results.xlsx", json=payload)

    assert response.status_code == 200
    assert response.content.startswith(b"PK")
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert response.headers["content-disposition"].endswith('.xlsx"')
