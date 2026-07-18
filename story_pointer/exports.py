"""Markdown/Excel-ready final-result export helpers.

The browser creates Markdown locally. Excel is produced server-side so the
download is a genuine XLSX workbook rather than a renamed CSV/HTML file.
"""
from __future__ import annotations

import io
from typing import Literal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field

from .schema import StoryPointResult


class ExportResultItem(BaseModel):
    index: int = Field(ge=0)
    title: str = Field(default="", max_length=1000)
    status: Literal["done", "failed"]
    result: StoryPointResult | None = None
    error: str = Field(default="", max_length=10_000)
    trace_id: str = Field(default="", max_length=256)


class ResultsExportRequest(BaseModel):
    items: list[ExportResultItem] = Field(min_length=1, max_length=1000)


def _cell(value: object) -> object:
    """Keep strings safe from spreadsheet-formula injection and XLSX limits."""
    if value is None:
        return ""
    if isinstance(value, (bool, int, float)):
        return value
    text = str(value)[:32_767]
    if text.lstrip().startswith(("=", "+", "-", "@")):
        text = "'" + text
    return text


def _append(sheet, values: list[object]) -> None:
    sheet.append([_cell(value) for value in values])


def _finish_sheet(sheet, widths: list[int]) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F6FEB")
        cell.alignment = Alignment(vertical="center")
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def build_results_workbook(items: list[ExportResultItem]) -> bytes:
    """Create a styled workbook with summary and normalized detail sheets."""
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    _append(summary, [
        "#", "Story", "Status", "Points", "TL;DR", "Why", "Person Days Min",
        "Person Days Max", "Must Split", "Spike Needed", "Provider", "Model", "Error", "Trace ID",
    ])

    factors = workbook.create_sheet("Factors")
    _append(factors, ["#", "Story", "Factor", "Level", "Evidence"])
    risks = workbook.create_sheet("Risks")
    _append(risks, ["#", "Story", "Severity", "Risk", "Mitigation"])
    supporting = workbook.create_sheet("Supporting detail")
    _append(supporting, ["#", "Story", "Category", "Item", "Detail", "Points/Level"])

    for item in items:
        result = item.result
        certified = bool(result and result.is_invariant_satisfied())
        status = "Certified" if certified else "Failed / redacted"
        pd = result.person_days if result else None
        _append(summary, [
            item.index + 1,
            item.title or (result.title if result else ""),
            status,
            result.points if certified else "",
            result.tldr if result else "",
            result.plain_language_why if result else "",
            pd.min if pd else "",
            pd.max if pd else "",
            result.must_split if result else False,
            result.spike_needed if result else False,
            result.provider if result else "",
            result.model if result else "",
            item.error or (result.error if result else ""),
            item.trace_id,
        ])
        if not result:
            continue

        title = item.title or result.title
        for factor in result.factors:
            _append(factors, [item.index + 1, title, factor.id, factor.level, factor.evidence])
        for risk in result.risks:
            _append(risks, [item.index + 1, title, risk.severity, risk.description, risk.mitigation])
        for driver in result.deciding_drivers:
            _append(supporting, [item.index + 1, title, "Deciding driver", driver.id, driver.why, ""])
        for anchor in result.closest_anchors:
            _append(supporting, [item.index + 1, title, "Closest anchor", anchor.points, anchor.why, anchor.points])
        for hidden in result.hidden_work:
            _append(supporting, [item.index + 1, title, "Hidden work", hidden, "", ""])
        for assumption in result.assumptions:
            _append(supporting, [item.index + 1, title, "Assumption", assumption, "", ""])
        for split in result.recommended_split:
            _append(supporting, [item.index + 1, title, "Recommended split", split.title, split.why, split.points])
        for layer, level in result.per_layer_effort.model_dump().items():
            _append(supporting, [item.index + 1, title, "Layer effort", layer, "", level])
        if result.spike_needed:
            _append(supporting, [item.index + 1, title, "Spike", result.spike_reason, "", ""])

    _finish_sheet(summary, [6, 36, 18, 10, 48, 60, 16, 16, 12, 13, 16, 26, 42, 34])
    _finish_sheet(factors, [6, 36, 28, 12, 60])
    _finish_sheet(risks, [6, 36, 12, 48, 48])
    _finish_sheet(supporting, [6, 36, 22, 42, 54, 16])

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


__all__ = ["ExportResultItem", "ResultsExportRequest", "build_results_workbook"]
